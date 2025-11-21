"""XLSX output generator

Generates Excel files with multiple sheets containing analytics data.
"""

from pathlib import Path
from typing import List, Dict
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_xlsx_file(
    recordings_data: List[Dict],
    daily_summary: Dict,
    hourly_data: Dict,
    word_freq: Counter,
    mode_data: Dict,
    topic_data: Dict,
    filler_data: Counter,
    bigram_freq: Counter,
    trigram_freq: Counter,
    sentence_summary: Dict,
    output_dir: Path,
):
    """Generate XLSX file with multiple sheets."""
    if not OPENPYXL_AVAILABLE:
        print("\n⚠ Skipping XLSX generation: openpyxl not installed")
        print("  Install with: pip install openpyxl")
        return

    print("\nGenerating XLSX file...")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Recordings Detail
    ws1 = wb.create_sheet("Recordings Detail")
    detail_fields = [
        "recording_id", "datetime", "date", "hour", "day_of_week",
        "duration_seconds", "duration_ms", "has_transcript", "word_count",
        "char_count", "words_per_minute", "filler_word_count", "filler_word_percentage",
        "sentence_count", "avg_words_per_sentence", "avg_chars_per_sentence",
        "mode_name", "model_name", "app_version", "processing_time_ms", "segment_count",
        "folder_name", "primary_topic", "secondary_topics"
    ]
    ws1.append(detail_fields)
    # Style header
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for rec in recordings_data:
        ws1.append([rec.get(field, "") for field in detail_fields])

    # Sheet 2: Daily Summary
    ws2 = wb.create_sheet("Daily Summary")
    ws2.append(["date", "recordings_count", "total_duration_seconds", "total_duration_hours",
                "total_words", "total_characters", "avg_duration_seconds", "avg_words_per_recording"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for date in sorted(daily_summary.keys()):
        data = daily_summary[date]
        ws2.append([
            data["date"], data["recordings_count"], data["total_duration_seconds"],
            data["total_duration_hours"], data["total_words"], data["total_characters"],
            data["avg_duration_seconds"], data["avg_words_per_recording"]
        ])

    # Sheet 3: Hourly Patterns
    ws3 = wb.create_sheet("Hourly Patterns")
    ws3.append(["hour", "recordings_count", "total_duration_seconds", "avg_duration_seconds"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for hour in sorted(hourly_data.keys()):
        data = hourly_data[hour]
        ws3.append([data["hour"], data["recordings_count"],
                    data["total_duration_seconds"], data["avg_duration_seconds"]])

    # Sheet 4: Word Frequency
    ws4 = wb.create_sheet("Word Frequency")
    ws4.append(["word", "frequency", "percentage"])
    for cell in ws4[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    total_words_count = sum(word_freq.values())
    for word, freq in word_freq.most_common(500):
        percentage = round((freq / total_words_count * 100), 4) if total_words_count > 0 else 0
        ws4.append([word, freq, percentage])

    # Sheet 5: Phrase Frequency
    ws5 = wb.create_sheet("Phrase Frequency")
    ws5.append(["phrase", "type", "frequency", "percentage"])
    for cell in ws5[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    total_bigrams = sum(bigram_freq.values())
    for phrase, freq in bigram_freq.most_common(100):
        percentage = round((freq / total_bigrams * 100), 4) if total_bigrams > 0 else 0
        ws5.append([phrase, "2-gram", freq, percentage])

    total_trigrams = sum(trigram_freq.values())
    for phrase, freq in trigram_freq.most_common(50):
        percentage = round((freq / total_trigrams * 100), 4) if total_trigrams > 0 else 0
        ws5.append([phrase, "3-gram", freq, percentage])

    # Sheet 6: Filler Word Analysis
    ws6 = wb.create_sheet("Filler Word Analysis")
    ws6.append(["filler_phrase", "count", "percentage"])
    for cell in ws6[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    total_fillers = sum(filler_data.values())
    for filler, count in filler_data.most_common():
        percentage = round((count / total_fillers * 100), 2) if total_fillers > 0 else 0
        ws6.append([filler, count, percentage])

    # Sheet 7: Sentence Metrics
    ws7 = wb.create_sheet("Sentence Metrics")
    ws7.append(list(sentence_summary.keys()))
    for cell in ws7[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws7.append(list(sentence_summary.values()))

    # Sheet 8: Mode Usage
    ws8 = wb.create_sheet("Mode Usage")
    ws8.append(["mode_name", "count", "percentage", "total_duration_seconds"])
    for cell in ws8[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for mode in sorted(mode_data.keys()):
        data = mode_data[mode]
        ws8.append([data["mode_name"], data["count"], data["percentage"], data["total_duration_seconds"]])

    # Sheet 9: Topic Distribution
    ws9 = wb.create_sheet("Topic Distribution")
    ws9.append(["topic", "recording_count", "total_duration_seconds",
                "percentage_of_recordings", "percentage_of_time"])
    for cell in ws9[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for topic in sorted(topic_data.keys(), key=lambda x: topic_data[x]["recording_count"], reverse=True):
        data = topic_data[topic]
        ws9.append([
            data["topic"], data["recording_count"], data["total_duration_seconds"],
            data["percentage_of_recordings"], data["percentage_of_time"]
        ])

    # Save workbook
    xlsx_file = output_dir / "analytics.xlsx"
    wb.save(xlsx_file)
    print(f"  ✓ {xlsx_file.name}")
