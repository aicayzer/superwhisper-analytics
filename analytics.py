#!/usr/bin/env python3
"""Generic analytics tool for Super Whisper recordings."""
import json
import csv
import wave
import re
import sys
import configparser
import argparse
from pathlib import Path
from collections import Counter
from datetime import datetime
from typing import Dict, List, Tuple, Optional

# Import from lib modules
from lib.processing.text_analysis import (
    clean_text,
    extract_words,
    extract_ngrams,
    count_filler_words,
    split_sentences,
    calculate_sentence_metrics
)
from lib.processing.validators import validate_date_format
from lib.processing.recording_processor import (
    parse_datetime,
    get_wav_duration,
    classify_topic,
    filter_by_date,
    process_recordings
)

# Try to import openpyxl for XLSX support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Topic keyword definitions
TOPIC_KEYWORDS = {
    "Code/Development": [
        "code", "repository", "repo", "github", "git", "commit", "function", "class",
        "python", "javascript", "typescript", "error", "bug", "fix", "implement",
        "script", "file", "module", "package", "import", "export", "test", "testing"
    ],
    "Documentation": [
        "documentation", "doc", "readme", "write", "document", "page", "section",
        "confluence", "notion", "markdown", "specification", "spec", "outline",
        "draft", "edit", "revise", "update", "content"
    ],
    "Data Engineering": [
        "bigquery", "sql", "dbt", "dataform", "query", "table", "dataset", "model",
        "staging", "mart", "intermediate", "transformation", "etl", "elt", "schema",
        "column", "field", "data", "analytics", "warehouse"
    ],
    "Project Management": [
        "ticket", "task", "todo", "epic", "story", "plan", "planning", "progress",
        "update", "meeting", "agenda", "action", "item", "deadline", "milestone",
        "sprint", "backlog", "priority"
    ],
    "Business Context": [
        "business", "banking", "customer", "lending", "savings", "account", "application",
        "product", "feature", "requirement", "stakeholder", "user", "client", "service"
    ],
    "Feedback/Instructions": [
        "please", "should", "need", "want", "think", "prefer", "suggest", "review",
        "check", "confirm", "update", "change", "modify", "improve", "better",
        "feedback", "instruction", "clarify", "understand"
    ],
    "Analysis": [
        "analyze", "analysis", "insight", "metric", "performance", "trend", "data",
        "report", "dashboard", "chart", "graph", "statistic", "measure", "evaluate",
        "review", "quarter", "q1", "q2", "q3", "q4"
    ],
    "Technical Architecture": [
        "architecture", "design", "system", "service", "api", "endpoint", "mcp",
        "server", "client", "framework", "tool", "tooling", "infrastructure",
        "deployment", "config", "configuration", "best practice", "standard"
    ]
}

# Common English stop words
STOP_WORDS = {
    "the", "a", "an", "and", "or", "but", "in", "on", "at", "to", "for", "of", "with",
    "by", "from", "as", "is", "was", "are", "were", "been", "be", "have", "has", "had",
    "do", "does", "did", "will", "would", "should", "could", "may", "might", "must",
    "can", "this", "that", "these", "those", "i", "you", "he", "she", "it", "we", "they",
    "me", "him", "her", "us", "them", "my", "your", "his", "her", "its", "our", "their",
    "what", "which", "who", "whom", "whose", "where", "when", "why", "how", "all", "each",
    "every", "both", "few", "more", "most", "other", "some", "such", "no", "nor", "not",
    "only", "own", "same", "so", "than", "too", "very", "just", "now"
}

# Filler words and phrases (multi-word phrases use regex patterns)
FILLER_WORDS = {
    # Single word fillers
    'um': r'\bum+\b',
    'uh': r'\buh+\b',
    'er': r'\ber+\b',
    'ah': r'\bah+\b',
    'like': r'\blike\b',
    'basically': r'\bbasically\b',
    'literally': r'\bliterally\b',
    'actually': r'\bactually\b',
    'honestly': r'\bhonestly\b',
    'seriously': r'\bseriously\b',
    'totally': r'\btotally\b',
    'right': r'\bright\b(?!\s+now)',  # Exclude "right now" which is meaningful
    'okay': r'\bokay\b',
    'ok': r'\bok\b',
    'well': r'\bwell\b(?=\s+[,.]|\s+$)',  # Only filler when at end or before punctuation
    'so': r'\bso\b(?=\s+[,.]|\s+$)',
    'yeah': r'\byeah\b',
    'yes': r'\byes\b(?=\s+[,.]|\s+$)',

    # Multi-word fillers
    'you know': r'\byou\s+know\b',
    'i mean': r'\bi\s+mean\b',
    'you see': r'\byou\s+see\b',
    'i think': r'\bi\s+think\b',
    'you think': r'\byou\s+think\b',
    'sort of': r'\bsort\s+of\b',
    'kind of': r'\bkind\s+of\b',
    'a bit': r'\ba\s+bit\b',
    'a little': r'\ba\s+little\b',
    'at the end of the day': r'\bat\s+the\s+end\s+of\s+the\s+day\b',
    'to be honest': r'\bto\s+be\s+honest\b',
    'if you will': r'\bif\s+you\s+will\b',
    'as it were': r'\bas\s+it\s+were\b',
}


def generate_csv_files(recordings_data: List[Dict], output_dir: Path):
    """Generate all CSV output files."""
    print("\nGenerating CSV files...")

    # 1. Recordings detail
    detail_file = output_dir / "recordings_detail.csv"
    detail_fields = [
        "recording_id", "datetime", "date", "hour", "day_of_week",
        "duration_seconds", "duration_ms", "has_transcript", "word_count",
        "char_count", "words_per_minute", "filler_word_count", "filler_word_percentage",
        "sentence_count", "avg_words_per_sentence", "avg_chars_per_sentence",
        "mode_name", "model_name", "app_version", "processing_time_ms", "segment_count",
        "folder_name", "primary_topic", "secondary_topics"
    ]

    with open(detail_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=detail_fields)
        writer.writeheader()
        for rec in recordings_data:
            writer.writerow({k: rec.get(k, "") for k in detail_fields})
    print(f"  ✓ {detail_file.name}")

    # 2. Daily summary
    daily_summary = {}
    for rec in recordings_data:
        date = rec["date"]
        if date not in daily_summary:
            daily_summary[date] = {
                "date": date,
                "recordings_count": 0,
                "total_duration_seconds": 0,
                "total_words": 0,
                "total_characters": 0
            }
        daily_summary[date]["recordings_count"] += 1
        daily_summary[date]["total_duration_seconds"] += rec["duration_seconds"]
        daily_summary[date]["total_words"] += rec["word_count"]
        daily_summary[date]["total_characters"] += rec["char_count"]

    daily_file = output_dir / "daily_summary.csv"
    with open(daily_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=[
            "date", "recordings_count", "total_duration_seconds",
            "total_duration_hours", "total_words", "total_characters",
            "avg_duration_seconds", "avg_words_per_recording"
        ])
        writer.writeheader()
        for date in sorted(daily_summary.keys()):
            data = daily_summary[date]
            data["total_duration_hours"] = round(data["total_duration_seconds"] / 3600, 2)
            data["avg_duration_seconds"] = round(
                data["total_duration_seconds"] / data["recordings_count"], 2
            ) if data["recordings_count"] > 0 else 0
            data["avg_words_per_recording"] = round(
                data["total_words"] / data["recordings_count"], 2
            ) if data["recordings_count"] > 0 else 0
            writer.writerow(data)
    print(f"  ✓ {daily_file.name}")

    # 3. Hourly patterns
    hourly_data = {}
    for rec in recordings_data:
        hour = rec["hour"]
        if hour not in hourly_data:
            hourly_data[hour] = {
                "hour": hour,
                "recordings_count": 0,
                "total_duration_seconds": 0
            }
        hourly_data[hour]["recordings_count"] += 1
        hourly_data[hour]["total_duration_seconds"] += rec["duration_seconds"]

    hourly_file = output_dir / "hourly_patterns.csv"
    with open(hourly_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=[
            "hour", "recordings_count", "total_duration_seconds", "avg_duration_seconds"
        ])
        writer.writeheader()
        for hour in sorted(hourly_data.keys()):
            data = hourly_data[hour]
            data["avg_duration_seconds"] = round(
                data["total_duration_seconds"] / data["recordings_count"], 2
            ) if data["recordings_count"] > 0 else 0
            writer.writerow(data)
    print(f"  ✓ {hourly_file.name}")

    # 4. Word frequency
    all_words = []
    for rec in recordings_data:
        if rec["transcript"]:
            all_words.extend(extract_words(rec["transcript"]))

    word_freq = Counter(all_words)
    total_words = sum(word_freq.values())

    word_file = output_dir / "word_frequency.csv"
    with open(word_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=["word", "frequency", "percentage"])
        writer.writeheader()
        for word, freq in word_freq.most_common(500):  # Top 500 words
            percentage = round((freq / total_words * 100), 4) if total_words > 0 else 0
            writer.writerow({"word": word, "frequency": freq, "percentage": percentage})
    print(f"  ✓ {word_file.name}")

    # 4b. Phrase frequency (n-grams)
    all_bigrams = []
    all_trigrams = []
    for rec in recordings_data:
        if rec["transcript"]:
            all_bigrams.extend(extract_ngrams(rec["transcript"], 2))
            all_trigrams.extend(extract_ngrams(rec["transcript"], 3))

    bigram_freq = Counter(all_bigrams)
    trigram_freq = Counter(all_trigrams)

    phrase_file = output_dir / "phrase_frequency.csv"
    with open(phrase_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=["phrase", "type", "frequency", "percentage"])
        writer.writeheader()

        # Write bigrams
        total_bigrams = sum(bigram_freq.values())
        for phrase, freq in bigram_freq.most_common(100):  # Top 100 bigrams
            percentage = round((freq / total_bigrams * 100), 4) if total_bigrams > 0 else 0
            writer.writerow({"phrase": phrase, "type": "2-gram", "frequency": freq, "percentage": percentage})

        # Write trigrams
        total_trigrams = sum(trigram_freq.values())
        for phrase, freq in trigram_freq.most_common(50):  # Top 50 trigrams
            percentage = round((freq / total_trigrams * 100), 4) if total_trigrams > 0 else 0
            writer.writerow({"phrase": phrase, "type": "3-gram", "frequency": freq, "percentage": percentage})

    print(f"  ✓ {phrase_file.name}")

    # 5. Mode usage
    mode_data = {}
    for rec in recordings_data:
        mode = rec["mode_name"]
        if mode not in mode_data:
            mode_data[mode] = {
                "mode_name": mode,
                "count": 0,
                "total_duration_seconds": 0
            }
        mode_data[mode]["count"] += 1
        mode_data[mode]["total_duration_seconds"] += rec["duration_seconds"]

    total_recordings = len(recordings_data)
    mode_file = output_dir / "mode_usage.csv"
    with open(mode_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=[
            "mode_name", "count", "percentage", "total_duration_seconds"
        ])
        writer.writeheader()
        for mode in sorted(mode_data.keys()):
            data = mode_data[mode]
            data["percentage"] = round((data["count"] / total_recordings * 100), 2) if total_recordings > 0 else 0
            writer.writerow(data)
    print(f"  ✓ {mode_file.name}")

    # 6. Topic distribution
    topic_data = {}
    for rec in recordings_data:
        topic = rec["primary_topic"]
        if topic not in topic_data:
            topic_data[topic] = {
                "topic": topic,
                "recording_count": 0,
                "total_duration_seconds": 0
            }
        topic_data[topic]["recording_count"] += 1
        topic_data[topic]["total_duration_seconds"] += rec["duration_seconds"]

    total_duration = sum(rec["duration_seconds"] for rec in recordings_data)
    topic_file = output_dir / "topic_distribution.csv"
    with open(topic_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=[
            "topic", "recording_count", "total_duration_seconds",
            "percentage_of_recordings", "percentage_of_time"
        ])
        writer.writeheader()
        for topic in sorted(topic_data.keys(), key=lambda x: topic_data[x]["recording_count"], reverse=True):
            data = topic_data[topic]
            data["percentage_of_recordings"] = round(
                (data["recording_count"] / total_recordings * 100), 2
            ) if total_recordings > 0 else 0
            data["percentage_of_time"] = round(
                (data["total_duration_seconds"] / total_duration * 100), 2
            ) if total_duration > 0 else 0
            writer.writerow(data)
    print(f"  ✓ {topic_file.name}")

    # 7. Filler word analysis
    all_fillers = Counter()
    for rec in recordings_data:
        if rec.get("filler_breakdown"):
            all_fillers.update(rec["filler_breakdown"])

    total_fillers = sum(all_fillers.values())
    filler_file = output_dir / "filler_word_analysis.csv"
    with open(filler_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=["filler_phrase", "count", "percentage"])
        writer.writeheader()
        for filler, count in all_fillers.most_common():
            percentage = round((count / total_fillers * 100), 2) if total_fillers > 0 else 0
            writer.writerow({"filler_phrase": filler, "count": count, "percentage": percentage})
    print(f"  ✓ {filler_file.name}")

    # 8. Sentence metrics summary
    recordings_with_sentences = [r for r in recordings_data if r.get("sentence_count", 0) > 0]

    if recordings_with_sentences:
        sentence_lengths = []
        words_per_sentence_all = []
        chars_per_sentence_all = []

        for rec in recordings_with_sentences:
            sentence_count = rec.get("sentence_count", 0)
            if sentence_count > 0:
                sentence_lengths.append(sentence_count)
                words_per_sentence_all.append(rec.get("avg_words_per_sentence", 0))
                chars_per_sentence_all.append(rec.get("avg_chars_per_sentence", 0))

        # Calculate aggregate statistics
        sentence_metrics_summary = {
            "total_recordings_analyzed": len(recordings_with_sentences),
            "total_sentences": sum(sentence_lengths),
            "avg_sentences_per_recording": round(sum(sentence_lengths) / len(sentence_lengths), 2) if sentence_lengths else 0,
            "min_sentences_per_recording": min(sentence_lengths) if sentence_lengths else 0,
            "max_sentences_per_recording": max(sentence_lengths) if sentence_lengths else 0,
            "avg_words_per_sentence": round(sum(words_per_sentence_all) / len(words_per_sentence_all), 2) if words_per_sentence_all else 0,
            "min_words_per_sentence": round(min(words_per_sentence_all), 2) if words_per_sentence_all else 0,
            "max_words_per_sentence": round(max(words_per_sentence_all), 2) if words_per_sentence_all else 0,
            "avg_chars_per_sentence": round(sum(chars_per_sentence_all) / len(chars_per_sentence_all), 2) if chars_per_sentence_all else 0,
        }
    else:
        sentence_metrics_summary = {
            "total_recordings_analyzed": 0,
            "total_sentences": 0,
            "avg_sentences_per_recording": 0,
            "min_sentences_per_recording": 0,
            "max_sentences_per_recording": 0,
            "avg_words_per_sentence": 0,
            "min_words_per_sentence": 0,
            "max_words_per_sentence": 0,
            "avg_chars_per_sentence": 0,
        }

    sentence_file = output_dir / "sentence_metrics.csv"
    with open(sentence_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=sentence_metrics_summary.keys())
        writer.writeheader()
        writer.writerow(sentence_metrics_summary)
    print(f"  ✓ {sentence_file.name}")

    return daily_summary, hourly_data, word_freq, mode_data, topic_data, all_fillers, bigram_freq, trigram_freq, sentence_metrics_summary


def generate_insights_report(recordings_data: List[Dict], output_dir: Path):
    """Generate markdown insights report."""
    print("\nGenerating insights report...")

    total_recordings = len(recordings_data)
    recordings_with_transcripts = sum(1 for r in recordings_data if r["has_transcript"])
    total_duration_seconds = sum(r["duration_seconds"] for r in recordings_data)
    total_duration_hours = total_duration_seconds / 3600
    total_words = sum(r["word_count"] for r in recordings_data)
    total_characters = sum(r["char_count"] for r in recordings_data)

    avg_duration = total_duration_seconds / total_recordings if total_recordings > 0 else 0
    avg_words = total_words / recordings_with_transcripts if recordings_with_transcripts > 0 else 0
    avg_wpm = sum(r["words_per_minute"] for r in recordings_data if r["duration_seconds"] > 0) / recordings_with_transcripts if recordings_with_transcripts > 0 else 0

    # Date range
    dates = sorted(set(r["date"] for r in recordings_data))
    first_date = dates[0] if dates else "Unknown"
    last_date = dates[-1] if dates else "Unknown"

    # Daily averages
    days_covered = len(dates)
    avg_recordings_per_day = total_recordings / days_covered if days_covered > 0 else 0
    avg_duration_per_day = total_duration_hours / days_covered if days_covered > 0 else 0

    # Hourly patterns
    hourly_counts = {}
    for r in recordings_data:
        hour = r["hour"]
        hourly_counts[hour] = hourly_counts.get(hour, 0) + 1
    peak_hour = max(hourly_counts.items(), key=lambda x: x[1])[0] if hourly_counts else None

    # Day of week patterns
    dow_counts = {}
    for r in recordings_data:
        dow = r["day_of_week"]
        dow_counts[dow] = dow_counts.get(dow, 0) + 1

    # Topic distribution
    topic_counts = {}
    for r in recordings_data:
        topic = r["primary_topic"]
        topic_counts[topic] = topic_counts.get(topic, 0) + 1

    # Mode distribution
    mode_counts = {}
    for r in recordings_data:
        mode = r["mode_name"]
        mode_counts[mode] = mode_counts.get(mode, 0) + 1

    # Longest recordings
    longest_by_duration = sorted(recordings_data, key=lambda x: x["duration_seconds"], reverse=True)[:5]
    longest_by_words = sorted([r for r in recordings_data if r["has_transcript"]], key=lambda x: x["word_count"], reverse=True)[:5]

    report = f"""# Super Whisper Recordings Analytics Report

Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

## Executive Summary

- **Total Recordings**: {total_recordings:,}
- **Recordings with Transcripts**: {recordings_with_transcripts:,} ({recordings_with_transcripts/total_recordings*100:.1f}%)
- **Total Recording Time**: {total_duration_hours:.1f} hours ({total_duration_seconds/60:.0f} minutes)
- **Total Words Transcribed**: {total_words:,}
- **Total Characters**: {total_characters:,}
- **Date Range**: {first_date} to {last_date} ({days_covered} days)
- **Average Recordings per Day**: {avg_recordings_per_day:.1f}
- **Average Duration per Day**: {avg_duration_per_day:.2f} hours

## Time Distribution Analysis

### Daily Patterns
- **Average Recordings per Day**: {avg_recordings_per_day:.1f}
- **Average Duration per Day**: {avg_duration_per_day:.2f} hours
- **Total Days Active**: {days_covered} days

### Hourly Patterns
- **Peak Hour**: {peak_hour}:00 ({hourly_counts.get(peak_hour, 0)} recordings)
- **Most Active Hours**: {', '.join([f"{h}:00 ({c})" for h, c in sorted(hourly_counts.items(), key=lambda x: x[1], reverse=True)[:5]])}

### Day of Week Patterns
"""

    for dow in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]:
        count = dow_counts.get(dow, 0)
        percentage = (count / total_recordings * 100) if total_recordings > 0 else 0
        report += f"- **{dow}**: {count} recordings ({percentage:.1f}%)\n"

    report += f"""
## Volume Metrics

- **Average Duration per Recording**: {avg_duration/60:.1f} minutes ({avg_duration:.1f} seconds)
- **Average Words per Recording**: {avg_words:.0f} words
- **Average Words per Minute**: {avg_wpm:.1f} WPM
- **Average Characters per Recording**: {total_characters/recordings_with_transcripts:.0f} characters

### Longest Recordings (by duration)
"""

    for i, rec in enumerate(longest_by_duration, 1):
        report += f"{i}. {rec['recording_id']}: {rec['duration_seconds']/60:.1f} minutes ({rec['word_count']} words) - {rec['primary_topic']}\n"

    report += f"""
### Longest Recordings (by word count)
"""

    for i, rec in enumerate(longest_by_words, 1):
        report += f"{i}. {rec['recording_id']}: {rec['word_count']} words ({rec['duration_seconds']/60:.1f} minutes) - {rec['primary_topic']}\n"

    report += f"""
## Topic Distribution

### Primary Topics
"""

    for topic, count in sorted(topic_counts.items(), key=lambda x: x[1], reverse=True):
        percentage = (count / total_recordings * 100) if total_recordings > 0 else 0
        report += f"- **{topic}**: {count} recordings ({percentage:.1f}%)\n"

    report += f"""
## Mode Usage

"""

    for mode, count in sorted(mode_counts.items(), key=lambda x: x[1], reverse=True):
        percentage = (count / total_recordings * 100) if total_recordings > 0 else 0
        report += f"- **{mode}**: {count} recordings ({percentage:.1f}%)\n"

    report += f"""
## Technical Metrics

- **Recordings with Transcripts**: {recordings_with_transcripts:,} ({recordings_with_transcripts/total_recordings*100:.1f}%)
- **Recordings without Transcripts**: {total_recordings - recordings_with_transcripts:,} ({(total_recordings - recordings_with_transcripts)/total_recordings*100:.1f}%)

### Processing Efficiency
"""

    recordings_with_processing = [r for r in recordings_data if r["processing_time_ms"] > 0]
    if recordings_with_processing:
        avg_processing_ratio = sum(
            r["processing_time_ms"] / r["duration_ms"]
            for r in recordings_with_processing if r["duration_ms"] > 0
        ) / len(recordings_with_processing)
        report += f"- **Average Processing Time Ratio**: {avg_processing_ratio*100:.2f}% (processing time / recording duration)\n"

    report += f"""
## Notable Patterns and Trends

### Recording Activity
- The dataset spans {days_covered} days with an average of {avg_recordings_per_day:.1f} recordings per day
- Peak activity occurs around {peak_hour}:00
- Most recordings are made during weekdays

### Content Patterns
- Average speech rate: {avg_wpm:.1f} words per minute
- {recordings_with_transcripts/total_recordings*100:.1f}% of recordings have transcripts
- Most common topic: {max(topic_counts.items(), key=lambda x: x[1])[0] if topic_counts else "Unknown"}

## Manual Analysis Section

*This section will be populated with deeper insights after manual review of the CSV data.*

### Key Observations
- Review the `recordings_detail.csv` file for detailed patterns
- Check `topic_distribution.csv` for topic trends over time
- Analyze `hourly_patterns.csv` for work schedule insights
- Examine `word_frequency.csv` for common terminology

### Suggested Further Analysis
1. Cross-reference topic distribution with time patterns
2. Identify correlations between recording length and topic
3. Analyze trends in mode usage over time
4. Review longest recordings for common themes
5. Examine word frequency for domain-specific terminology

---

*Report generated from {total_recordings:,} recordings*
"""

    report_file = output_dir / "insights_report.md"
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f"  ✓ {report_file.name}")


def generate_xlsx_file(recordings_data: List[Dict], daily_summary: Dict, hourly_data: Dict,
                       word_freq: Counter, mode_data: Dict, topic_data: Dict, filler_data: Counter,
                       bigram_freq: Counter, trigram_freq: Counter, sentence_summary: Dict, output_dir: Path):
    """Generate XLSX file with multiple sheets."""
    if not OPENPYXL_AVAILABLE:
        print("\n⚠ Skipping XLSX generation: openpyxl not installed")
        print("  Install with: pip install openpyxl")
        return

    print("\nGenerating XLSX file...")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Recordings Detail
    ws1 = wb.create_sheet("Recordings Detail")
    detail_fields = [
        "recording_id", "datetime", "date", "hour", "day_of_week",
        "duration_seconds", "duration_ms", "has_transcript", "word_count",
        "char_count", "words_per_minute", "filler_word_count", "filler_word_percentage",
        "sentence_count", "avg_words_per_sentence", "avg_chars_per_sentence",
        "mode_name", "model_name", "app_version", "processing_time_ms", "segment_count",
        "folder_name", "primary_topic", "secondary_topics"
    ]
    ws1.append(detail_fields)
    # Style header
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for rec in recordings_data:
        ws1.append([rec.get(k, "") for k in detail_fields])

    # Sheet 2: Daily Summary
    ws2 = wb.create_sheet("Daily Summary")
    ws2.append(["date", "recordings_count", "total_duration_seconds",
                "total_duration_hours", "total_words", "total_characters",
                "avg_duration_seconds", "avg_words_per_recording"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for date in sorted(daily_summary.keys()):
        data = daily_summary[date]
        data["total_duration_hours"] = round(data["total_duration_seconds"] / 3600, 2)
        data["avg_duration_seconds"] = round(
            data["total_duration_seconds"] / data["recordings_count"], 2
        ) if data["recordings_count"] > 0 else 0
        data["avg_words_per_recording"] = round(
            data["total_words"] / data["recordings_count"], 2
        ) if data["recordings_count"] > 0 else 0
        ws2.append([
            data["date"], data["recordings_count"], data["total_duration_seconds"],
            data["total_duration_hours"], data["total_words"], data["total_characters"],
            data["avg_duration_seconds"], data["avg_words_per_recording"]
        ])

    # Sheet 3: Hourly Patterns
    ws3 = wb.create_sheet("Hourly Patterns")
    ws3.append(["hour", "recordings_count", "total_duration_seconds", "avg_duration_seconds"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for hour in sorted(hourly_data.keys()):
        data = hourly_data[hour]
        data["avg_duration_seconds"] = round(
            data["total_duration_seconds"] / data["recordings_count"], 2
        ) if data["recordings_count"] > 0 else 0
        ws3.append([data["hour"], data["recordings_count"],
                   data["total_duration_seconds"], data["avg_duration_seconds"]])

    # Sheet 4: Word Frequency
    ws4 = wb.create_sheet("Word Frequency")
    ws4.append(["word", "frequency", "percentage"])
    for cell in ws4[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    total_words = sum(word_freq.values())
    for word, freq in word_freq.most_common(500):
        percentage = round((freq / total_words * 100), 4) if total_words > 0 else 0
        ws4.append([word, freq, percentage])

    # Sheet 5: Mode Usage
    ws5 = wb.create_sheet("Mode Usage")
    ws5.append(["mode_name", "count", "percentage", "total_duration_seconds"])
    for cell in ws5[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    total_recordings = len(recordings_data)
    for mode in sorted(mode_data.keys()):
        data = mode_data[mode]
        data["percentage"] = round((data["count"] / total_recordings * 100), 2) if total_recordings > 0 else 0
        ws5.append([data["mode_name"], data["count"], data["percentage"], data["total_duration_seconds"]])

    # Sheet 6: Topic Distribution
    ws6 = wb.create_sheet("Topic Distribution")
    ws6.append(["topic", "recording_count", "total_duration_seconds",
                "percentage_of_recordings", "percentage_of_time"])
    for cell in ws6[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    total_duration = sum(rec["duration_seconds"] for rec in recordings_data)
    for topic in sorted(topic_data.keys(), key=lambda x: topic_data[x]["recording_count"], reverse=True):
        data = topic_data[topic]
        data["percentage_of_recordings"] = round(
            (data["recording_count"] / total_recordings * 100), 2
        ) if total_recordings > 0 else 0
        data["percentage_of_time"] = round(
            (data["total_duration_seconds"] / total_duration * 100), 2
        ) if total_duration > 0 else 0
        ws6.append([
            data["topic"], data["recording_count"], data["total_duration_seconds"],
            data["percentage_of_recordings"], data["percentage_of_time"]
        ])

    # Sheet 7: Filler Word Analysis
    ws7 = wb.create_sheet("Filler Word Analysis")
    ws7.append(["filler_phrase", "count", "percentage"])
    for cell in ws7[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    total_fillers = sum(filler_data.values())
    for filler, count in filler_data.most_common():
        percentage = round((count / total_fillers * 100), 2) if total_fillers > 0 else 0
        ws7.append([filler, count, percentage])

    # Sheet 8: Phrase Frequency
    ws8 = wb.create_sheet("Phrase Frequency")
    ws8.append(["phrase", "type", "frequency", "percentage"])
    for cell in ws8[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Add bigrams
    total_bigrams = sum(bigram_freq.values())
    for phrase, freq in bigram_freq.most_common(100):
        percentage = round((freq / total_bigrams * 100), 4) if total_bigrams > 0 else 0
        ws8.append([phrase, "2-gram", freq, percentage])

    # Add trigrams
    total_trigrams = sum(trigram_freq.values())
    for phrase, freq in trigram_freq.most_common(50):
        percentage = round((freq / total_trigrams * 100), 4) if total_trigrams > 0 else 0
        ws8.append([phrase, "3-gram", freq, percentage])

    # Sheet 9: Sentence Metrics
    ws9 = wb.create_sheet("Sentence Metrics")
    ws9.append(list(sentence_summary.keys()))
    for cell in ws9[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws9.append(list(sentence_summary.values()))

    xlsx_file = output_dir / "analytics.xlsx"
    wb.save(xlsx_file)
    print(f"  ✓ {xlsx_file.name}")


def generate_json_file(recordings_data: List[Dict], daily_summary: Dict, hourly_data: Dict,
                       word_freq: Counter, mode_data: Dict, topic_data: Dict, filler_data: Counter,
                       bigram_freq: Counter, trigram_freq: Counter, sentence_summary: Dict, output_dir: Path):
    """Generate JSON file with structured data."""
    print("\nGenerating JSON file...")

    total_recordings = len(recordings_data)
    recordings_with_transcripts = sum(1 for r in recordings_data if r["has_transcript"])
    total_duration_seconds = sum(r["duration_seconds"] for r in recordings_data)
    total_words = sum(r["word_count"] for r in recordings_data)
    total_characters = sum(r["char_count"] for r in recordings_data)

    dates = sorted(set(r["date"] for r in recordings_data))
    first_date = dates[0] if dates else "Unknown"
    last_date = dates[-1] if dates else "Unknown"

    # Prepare recordings without transcript and filler breakdown fields
    recordings_clean = []
    for rec in recordings_data:
        rec_copy = rec.copy()
        rec_copy.pop("transcript", None)  # Remove transcript from JSON
        rec_copy.pop("filler_breakdown", None)  # Remove detailed breakdown
        recordings_clean.append(rec_copy)

    # Prepare daily summary
    daily_summary_list = []
    for date in sorted(daily_summary.keys()):
        data = daily_summary[date].copy()
        data["total_duration_hours"] = round(data["total_duration_seconds"] / 3600, 2)
        data["avg_duration_seconds"] = round(
            data["total_duration_seconds"] / data["recordings_count"], 2
        ) if data["recordings_count"] > 0 else 0
        data["avg_words_per_recording"] = round(
            data["total_words"] / data["recordings_count"], 2
        ) if data["recordings_count"] > 0 else 0
        daily_summary_list.append(data)

    # Prepare hourly patterns
    hourly_patterns_list = []
    for hour in sorted(hourly_data.keys()):
        data = hourly_data[hour].copy()
        data["avg_duration_seconds"] = round(
            data["total_duration_seconds"] / data["recordings_count"], 2
        ) if data["recordings_count"] > 0 else 0
        hourly_patterns_list.append(data)

    # Prepare word frequency
    total_words_count = sum(word_freq.values())
    word_frequency_list = [
        {"word": word, "frequency": freq, "percentage": round((freq / total_words_count * 100), 4)}
        for word, freq in word_freq.most_common(500)
    ]

    # Prepare mode usage
    mode_usage_list = []
    for mode in sorted(mode_data.keys()):
        data = mode_data[mode].copy()
        data["percentage"] = round((data["count"] / total_recordings * 100), 2) if total_recordings > 0 else 0
        mode_usage_list.append(data)

    # Prepare topic distribution
    total_duration = sum(rec["duration_seconds"] for rec in recordings_data)
    topic_distribution_list = []
    for topic in sorted(topic_data.keys(), key=lambda x: topic_data[x]["recording_count"], reverse=True):
        data = topic_data[topic].copy()
        data["percentage_of_recordings"] = round(
            (data["recording_count"] / total_recordings * 100), 2
        ) if total_recordings > 0 else 0
        data["percentage_of_time"] = round(
            (data["total_duration_seconds"] / total_duration * 100), 2
        ) if total_duration > 0 else 0
        topic_distribution_list.append(data)

    # Prepare filler word analysis
    total_fillers_count = sum(filler_data.values())
    filler_analysis_list = [
        {"filler_phrase": filler, "count": count, "percentage": round((count / total_fillers_count * 100), 2)}
        for filler, count in filler_data.most_common()
    ]

    # Prepare phrase frequency
    total_bigrams_count = sum(bigram_freq.values())
    total_trigrams_count = sum(trigram_freq.values())
    phrase_frequency_list = []

    # Add bigrams
    for phrase, count in bigram_freq.most_common(100):
        phrase_frequency_list.append({
            "phrase": phrase,
            "type": "2-gram",
            "frequency": count,
            "percentage": round((count / total_bigrams_count * 100), 4) if total_bigrams_count > 0 else 0
        })

    # Add trigrams
    for phrase, count in trigram_freq.most_common(50):
        phrase_frequency_list.append({
            "phrase": phrase,
            "type": "3-gram",
            "frequency": count,
            "percentage": round((count / total_trigrams_count * 100), 4) if total_trigrams_count > 0 else 0
        })

    json_data = {
        "metadata": {
            "generated_at": datetime.now().isoformat(),
            "total_recordings": total_recordings,
            "recordings_with_transcripts": recordings_with_transcripts,
            "total_duration_seconds": round(total_duration_seconds, 2),
            "total_duration_hours": round(total_duration_seconds / 3600, 2),
            "total_words": total_words,
            "total_characters": total_characters,
            "date_range": {
                "first_date": first_date,
                "last_date": last_date,
                "days_covered": len(dates)
            }
        },
        "summary": {
            "avg_recordings_per_day": round(total_recordings / len(dates), 2) if dates else 0,
            "avg_duration_per_day_hours": round((total_duration_seconds / 3600) / len(dates), 2) if dates else 0,
            "avg_words_per_recording": round(total_words / recordings_with_transcripts, 2) if recordings_with_transcripts > 0 else 0
        },
        "recordings": recordings_clean,
        "daily_summary": daily_summary_list,
        "hourly_patterns": hourly_patterns_list,
        "word_frequency": word_frequency_list,
        "mode_usage": mode_usage_list,
        "topic_distribution": topic_distribution_list,
        "filler_word_analysis": filler_analysis_list,
        "phrase_frequency": phrase_frequency_list,
        "sentence_metrics": sentence_summary
    }

    json_file = output_dir / "analytics.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False)
    print(f"  ✓ {json_file.name}")


def load_config() -> configparser.ConfigParser:
    """Load configuration from config files with fallback."""
    config = configparser.ConfigParser()
    script_dir = Path(__file__).parent

    # Set defaults
    config['paths'] = {
        'recordings_dir': '../recordings',
        'output_dir': './outputs'
    }
    config['analysis'] = {
        'default_top_words': '500',
        'default_top_bigrams': '100',
        'default_top_trigrams': '50'
    }

    # Try to load config.ini first
    config_file = script_dir / 'config.ini'
    if config_file.exists():
        config.read(config_file)

    # Override with local config if it exists
    local_config_file = script_dir / 'config.local.ini'
    if local_config_file.exists():
        config.read(local_config_file)

    return config


def resolve_path(path_str: str, script_dir: Path) -> Path:
    """Resolve a path string to an absolute Path, handling relative paths."""
    path = Path(path_str)
    if path.is_absolute():
        return path
    else:
        return (script_dir / path).resolve()


def escape_mermaid_label(text: str) -> str:
    """Escape special characters for Mermaid chart labels."""
    # Replace problematic characters
    text = str(text).replace('"', "'").replace('\n', ' ').replace('\r', ' ')
    # Limit length to avoid chart rendering issues
    if len(text) > 50:
        text = text[:47] + "..."
    return text


def generate_mermaid_charts(recordings_data: List[Dict], daily_summary: Dict,
                            word_freq: Counter, mode_data: Dict, topic_data: Dict, output_dir: Path):
    """Generate Mermaid chart files for visualizations."""
    print("\nGenerating Mermaid charts...")

    # 1. Daily Activity Timeline
    if daily_summary:
        dates = sorted(daily_summary.keys())
        days_count = len(dates)

        # Auto-aggregate to weekly if more than 30 days
        if days_count > 30:
            # Aggregate by week
            from datetime import datetime as dt_module
            weekly_data = {}
            for date_str in dates:
                date_obj = dt_module.fromisoformat(date_str)
                week_key = date_obj.strftime("%Y-W%W")  # Year-WeekNumber
                if week_key not in weekly_data:
                    weekly_data[week_key] = {"recordings": 0, "duration": 0}
                weekly_data[week_key]["recordings"] += daily_summary[date_str]["recordings_count"]
                weekly_data[week_key]["duration"] += daily_summary[date_str]["total_duration_seconds"] / 3600

            timeline_chart = "```mermaid\n---\nconfig:\n  xyChart:\n    width: 900\n    height: 400\n---\nxyChart-beta\n"
            timeline_chart += "    title \"Weekly Recording Activity\"\n"
            timeline_chart += "    x-axis [" + ", ".join([f'"{escape_mermaid_label(w)}"' for w in sorted(weekly_data.keys())]) + "]\n"
            timeline_chart += "    y-axis \"Recordings Count\" 0 --> " + str(max(w["recordings"] for w in weekly_data.values()) + 5) + "\n"
            timeline_chart += "    line [" + ", ".join([str(weekly_data[w]["recordings"]) for w in sorted(weekly_data.keys())]) + "]\n"
            timeline_chart += "```\n"
        else:
            # Daily timeline
            timeline_chart = "```mermaid\n---\nconfig:\n  xyChart:\n    width: 900\n    height: 400\n---\nxyChart-beta\n"
            timeline_chart += "    title \"Daily Recording Activity\"\n"
            timeline_chart += "    x-axis [" + ", ".join([f'"{escape_mermaid_label(d)}"' for d in dates]) + "]\n"
            max_count = max(daily_summary[d]["recordings_count"] for d in dates)
            timeline_chart += "    y-axis \"Recordings Count\" 0 --> " + str(max_count + 5) + "\n"
            timeline_chart += "    line [" + ", ".join([str(daily_summary[d]["recordings_count"]) for d in dates]) + "]\n"
            timeline_chart += "```\n"

        timeline_file = output_dir / "timeline_activity.mmd"
        with open(timeline_file, 'w', encoding='utf-8') as f:
            f.write(timeline_chart)
        print(f"  ✓ {timeline_file.name}")

    # 2. Topic Distribution Over Time
    if recordings_data and len(recordings_data) > 10:
        # Group by date and topic
        from datetime import datetime as dt_module
        from collections import defaultdict

        date_topic_counts = defaultdict(lambda: defaultdict(int))
        for rec in recordings_data:
            date = rec["date"]
            topic = rec["primary_topic"]
            date_topic_counts[date][topic] += 1

        dates = sorted(date_topic_counts.keys())
        all_topics = set()
        for topics_dict in date_topic_counts.values():
            all_topics.update(topics_dict.keys())

        # Get top 5 topics by total count
        topic_totals = Counter()
        for topics_dict in date_topic_counts.values():
            topic_totals.update(topics_dict)
        top_topics = [t for t, _ in topic_totals.most_common(5)]

        # Aggregate if more than 30 days
        if len(dates) > 30:
            weekly_topic_data = defaultdict(lambda: defaultdict(int))
            for date_str in dates:
                date_obj = dt_module.fromisoformat(date_str)
                week_key = date_obj.strftime("%Y-W%W")
                for topic, count in date_topic_counts[date_str].items():
                    if topic in top_topics:
                        weekly_topic_data[week_key][topic] += count

            weeks = sorted(weekly_topic_data.keys())
            topic_timeline = "```mermaid\n---\nconfig:\n  xyChart:\n    width: 900\n    height: 500\n---\nxyChart-beta\n"
            topic_timeline += "    title \"Weekly Topic Distribution\"\n"
            topic_timeline += "    x-axis [" + ", ".join([f'"{escape_mermaid_label(w)}"' for w in weeks]) + "]\n"
            topic_timeline += "    y-axis \"Recording Count\"\n"

            for topic in top_topics:
                counts = [weekly_topic_data[w].get(topic, 0) for w in weeks]
                topic_timeline += f"    line \"{escape_mermaid_label(topic)}\" [" + ", ".join([str(c) for c in counts]) + "]\n"
            topic_timeline += "```\n"
        else:
            topic_timeline = "```mermaid\n---\nconfig:\n  xyChart:\n    width: 900\n    height: 500\n---\nxyChart-beta\n"
            topic_timeline += "    title \"Daily Topic Distribution (Top 5)\"\n"
            topic_timeline += "    x-axis [" + ", ".join([f'"{escape_mermaid_label(d)}"' for d in dates]) + "]\n"
            topic_timeline += "    y-axis \"Recording Count\"\n"

            for topic in top_topics:
                counts = [date_topic_counts[d].get(topic, 0) for d in dates]
                topic_timeline += f"    line \"{escape_mermaid_label(topic)}\" [" + ", ".join([str(c) for c in counts]) + "]\n"
            topic_timeline += "```\n"

        topic_timeline_file = output_dir / "timeline_topics.mmd"
        with open(topic_timeline_file, 'w', encoding='utf-8') as f:
            f.write(topic_timeline)
        print(f"  ✓ {topic_timeline_file.name}")

    # 3. Top Words Bar Chart
    if word_freq:
        top_words = word_freq.most_common(20)
        words_chart = "```mermaid\n---\nconfig:\n  xyChart:\n    width: 900\n    height: 600\n---\nxyChart-beta horizontal\n"
        words_chart += "    title \"Top 20 Most Common Words\"\n"
        words_chart += "    x-axis \"Frequency\"\n"
        words_chart += "    y-axis [" + ", ".join([f'"{escape_mermaid_label(w)}"' for w, _ in top_words]) + "]\n"
        words_chart += "    bar [" + ", ".join([str(c) for _, c in top_words]) + "]\n"
        words_chart += "```\n"

        words_file = output_dir / "chart_top_words.mmd"
        with open(words_file, 'w', encoding='utf-8') as f:
            f.write(words_chart)
        print(f"  ✓ {words_file.name}")

    # 4. Mode Usage Bar Chart
    if mode_data:
        sorted_modes = sorted(mode_data.items(), key=lambda x: x[1]["count"], reverse=True)
        mode_chart = "```mermaid\n---\nconfig:\n  xyChart:\n    width: 900\n    height: 400\n---\nxyChart-beta horizontal\n"
        mode_chart += "    title \"Recording Mode Usage\"\n"
        mode_chart += "    x-axis \"Number of Recordings\"\n"
        mode_chart += "    y-axis [" + ", ".join([f'"{escape_mermaid_label(mode)}"' for mode, _ in sorted_modes]) + "]\n"
        mode_chart += "    bar [" + ", ".join([str(data["count"]) for _, data in sorted_modes]) + "]\n"
        mode_chart += "```\n"

        mode_file = output_dir / "chart_mode_usage.mmd"
        with open(mode_file, 'w', encoding='utf-8') as f:
            f.write(mode_chart)
        print(f"  ✓ {mode_file.name}")

    # 5. Topic Distribution Bar Chart
    if topic_data:
        sorted_topics = sorted(topic_data.items(), key=lambda x: x[1]["recording_count"], reverse=True)
        topic_chart = "```mermaid\n---\nconfig:\n  xyChart:\n    width: 900\n    height: 500\n---\nxyChart-beta horizontal\n"
        topic_chart += "    title \"Topic Distribution\"\n"
        topic_chart += "    x-axis \"Number of Recordings\"\n"
        topic_chart += "    y-axis [" + ", ".join([f'"{escape_mermaid_label(topic)}"' for topic, _ in sorted_topics]) + "]\n"
        topic_chart += "    bar [" + ", ".join([str(data["recording_count"]) for _, data in sorted_topics]) + "]\n"
        topic_chart += "```\n"

        topic_file = output_dir / "chart_topic_distribution.mmd"
        with open(topic_file, 'w', encoding='utf-8') as f:
            f.write(topic_chart)
        print(f"  ✓ {topic_file.name}")


def generate_ai_prompt_file(output_dir: Path):
    """Generate AI prompt file for insights generation."""
    print("\nGenerating AI prompt file...")

    prompt = """# AI Insights Generation Prompt

You are analyzing Super Whisper recording analytics data. The following files are available in this directory:

## Available Data Files

### Structured Data
1. **analytics.json** - Complete structured data with all metrics including new text analysis
2. **analytics.xlsx** - Excel workbook with multiple sheets:
   - Recordings Detail (with filler words and sentence metrics)
   - Daily Summary
   - Hourly Patterns
   - Word Frequency
   - Phrase Frequency (bigrams and trigrams)
   - Filler Word Analysis
   - Sentence Metrics
   - Mode Usage
   - Topic Distribution

### CSV Files
3. **recordings_detail.csv** - Full detailed data for each recording
4. **daily_summary.csv** - Aggregated daily statistics
5. **hourly_patterns.csv** - Activity patterns by hour
6. **word_frequency.csv** - Most common words (top 500)
7. **phrase_frequency.csv** - Common 2-grams and 3-grams (top 150)
8. **filler_word_analysis.csv** - Filler word/phrase usage breakdown
9. **sentence_metrics.csv** - Aggregate sentence-level statistics
10. **mode_usage.csv** - Distribution across recording modes
11. **topic_distribution.csv** - Topic classification statistics

### Visualisations
12. **timeline_activity.mmd** - Daily/weekly recording activity timeline (Mermaid)
13. **timeline_topics.mmd** - Topic distribution over time (Mermaid)
14. **chart_top_words.mmd** - Top 20 words bar chart (Mermaid)
15. **chart_mode_usage.mmd** - Mode usage bar chart (Mermaid)
16. **chart_topic_distribution.mmd** - Topic distribution bar chart (Mermaid)

### Reports
17. **insights_report.md** - Basic summary report (may need enhancement)

## Analysis Tasks

Please analyze the data and provide insights on:

### 1. Time Patterns
- Identify peak activity periods (hours, days of week)
- Analyze trends over time (increasing/decreasing activity)
- Identify any anomalies or unusual patterns
- Correlate activity with day of week

### 2. Content Analysis
- Review word frequency for domain-specific terminology
- Analyze phrase patterns (bigrams/trigrams) for common expressions
- Identify filler word usage patterns and speaking habits
- Examine sentence structure (length, complexity)
- Identify common themes from top words and phrases
- Analyze topic distribution patterns
- Look for correlations between topics and time patterns

### 3. Usage Patterns
- Analyze mode usage trends
- Identify most common use cases
- Review recording length distributions
- Examine speech rate patterns
- Compare filler word usage across different contexts

### 4. Communication Insights
- Analyze sentence structure trends (simple vs complex)
- Identify speaking patterns from filler word analysis
- Review phrase usage for common expressions or jargon
- Examine how communication style varies by topic or time

### 5. Insights Generation
- Provide actionable insights
- Identify interesting patterns or anomalies
- Suggest areas for further analysis
- Highlight key findings
- Note any communication improvements (filler word reduction, clearer sentences)
- Identify productivity patterns from timeline charts

## Output Format

Please generate an enhanced insights report that includes:
- Executive summary with key findings
- Detailed analysis sections
- Visualizations suggestions (if applicable)
- Recommendations or observations
- Any notable patterns or trends

## Data Structure Notes

- Recordings are identified by Unix timestamp folder names
- Duration is in seconds (duration_ms / 1000)
- Topics are classified using keyword matching (8 categories)
- Word frequency excludes common stop words (70+ words)
- Phrase frequency includes bigrams (2-word) and trigrams (3-word)
- Filler words include 30+ patterns (single and multi-word phrases)
- Sentence metrics calculated with abbreviation handling
- All dates are in ISO format (YYYY-MM-DD)
- Mermaid charts use xyChart-beta syntax
- Timeline charts auto-aggregate to weekly when >30 days

## Instructions

1. Load and examine the JSON file first for overall structure
2. Use CSV files for detailed analysis or specific queries
3. Review Mermaid charts for visual patterns (view .mmd files in Markdown viewer)
4. Cross-reference data across different files
5. Generate comprehensive insights beyond what's in the basic report
6. Be specific with numbers and percentages
7. Identify trends and patterns
8. Analyze text quality metrics (filler words, sentence structure)
9. Look for correlations between different metrics
10. Provide context-aware analysis

## New Analytics Features

This analysis includes enhanced text analysis:
- **Filler Word Detection**: Identifies verbal fillers like "um", "uh", "you know", "I mean"
- **Phrase Analysis**: Common 2-word and 3-word expressions
- **Sentence Metrics**: Average sentence length, complexity indicators
- **Timeline Visualisations**: Activity and topic trends over time

Consider these new dimensions when generating insights.

Begin your analysis now.
"""

    prompt_file = output_dir / "insights_prompt.md"
    with open(prompt_file, 'w', encoding='utf-8') as f:
        f.write(prompt)
    print(f"  ✓ {prompt_file.name}")


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Analyze Super Whisper recordings with optional date filtering.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process all recordings
  python3 analytics.py

  # Filter by specific date
  python3 analytics.py --date 2025-01-15

  # Filter by month
  python3 analytics.py --month 2025-01

  # Filter by date range
  python3 analytics.py --date-from 2025-01-01 --date-to 2025-01-31

  # Combine month and date range
  python3 analytics.py --month 2025-01 --date-from 2025-01-15
        """
    )

    parser.add_argument(
        '--date',
        type=str,
        help='Filter recordings by specific date (YYYY-MM-DD format)'
    )
    parser.add_argument(
        '--month',
        type=str,
        help='Filter recordings by month (YYYY-MM format)'
    )
    parser.add_argument(
        '--date-from',
        type=str,
        help='Filter recordings from this date onwards (YYYY-MM-DD format)'
    )
    parser.add_argument(
        '--date-to',
        type=str,
        help='Filter recordings up to this date (YYYY-MM-DD format)'
    )

    return parser.parse_args()


def main():
    """Main execution function."""
    # Parse command line arguments
    args = parse_arguments()

    # Validate date formats
    validate_date_format(args.date, 'date')
    validate_date_format(args.month, 'month')
    validate_date_format(args.date_from, 'date')
    validate_date_format(args.date_to, 'date')

    script_dir = Path(__file__).parent

    # Load configuration
    config = load_config()

    # Resolve paths from configuration
    recordings_dir = resolve_path(config['paths']['recordings_dir'], script_dir)
    outputs_base = resolve_path(config['paths']['output_dir'], script_dir)

    # Validate recordings directory
    if not recordings_dir.exists():
        print(f"Error: Recordings directory does not exist: {recordings_dir}")
        print(f"\nPlease check your configuration:")
        print(f"  - config.local.ini (if it exists)")
        print(f"  - config.ini")
        print(f"\nExpected recordings directory at: {recordings_dir}")
        sys.exit(1)

    if not recordings_dir.is_dir():
        print(f"Error: Recordings path is not a directory: {recordings_dir}")
        sys.exit(1)

    # Create timestamped output folder
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    outputs_base.mkdir(exist_ok=True)
    output_dir = outputs_base / timestamp
    output_dir.mkdir(exist_ok=True)

    print("=" * 60)
    print("Super Whisper Recordings Analytics")
    print("=" * 60)
    print(f"Recordings directory: {recordings_dir}")
    print(f"Output directory: {output_dir}")

    # Display active filters
    if args.date or args.month or args.date_from or args.date_to:
        print("\nActive filters:")
        if args.date:
            print(f"  - Date: {args.date}")
        if args.month:
            print(f"  - Month: {args.month}")
        if args.date_from:
            print(f"  - From: {args.date_from}")
        if args.date_to:
            print(f"  - To: {args.date_to}")
        print()

    # Process recordings
    recordings_data = process_recordings(
        recordings_dir,
        date_filter=args.date,
        month_filter=args.month,
        date_from=args.date_from,
        date_to=args.date_to
    )

    if not recordings_data:
        print("\nNo recordings found matching the specified criteria!")
        if args.date or args.month or args.date_from or args.date_to:
            print("Try adjusting your date filters or removing them to see all recordings.")
        sys.exit(1)

    # Generate CSV files (returns aggregated data)
    daily_summary, hourly_data, word_freq, mode_data, topic_data, filler_data, bigram_freq, trigram_freq, sentence_summary = generate_csv_files(recordings_data, output_dir)

    # Generate insights report
    generate_insights_report(recordings_data, output_dir)

    # Generate XLSX file
    generate_xlsx_file(recordings_data, daily_summary, hourly_data, word_freq, mode_data, topic_data, filler_data, bigram_freq, trigram_freq, sentence_summary, output_dir)

    # Generate JSON file
    generate_json_file(recordings_data, daily_summary, hourly_data, word_freq, mode_data, topic_data, filler_data, bigram_freq, trigram_freq, sentence_summary, output_dir)

    # Generate Mermaid charts
    generate_mermaid_charts(recordings_data, daily_summary, word_freq, mode_data, topic_data, output_dir)

    # Generate AI prompt file
    generate_ai_prompt_file(output_dir)

    print("\n" + "=" * 60)
    print("Analytics generation complete!")
    print(f"Output files saved to: {output_dir}")
    print("=" * 60)


if __name__ == "__main__":
    main()

